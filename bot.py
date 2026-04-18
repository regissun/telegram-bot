# bot.py
import os
import io
import re
import time
import json
import requests
import pandas as pd
from datetime import datetime
from flask import Flask, request, abort, jsonify
from openpyxl import Workbook, load_workbook

# ====== Cấu hình ======
LOG_FILE = "bot_user_log.xlsx"
OUTLOOK_LINK = os.getenv("OUTLOOK_LINK", "https://1drv.ms/x/c/63897167e619733d/EQCzDiktLplDjqgol9KBuZEBs0yxrsbyjxw49RRTE6uCuA?e=q1QOId&nav=MTVfezAwMDAwMDAwLTAwMDEtMDAwMC0wMDAwLTAwMDAwMDAwMDAwMH0")
TACT_LINK = os.getenv("TACT_LINK", "https://docs.google.com/spreadsheets/d/e/2PACX-1vSxJMSJZcwlD4ZUiY0a_N1KfeAyKp2HDUGzhXWA1wDxRkU1fFCU3BjfQZnquOEtwA/pubhtml?gid=248455740&single=true")

TOKEN = os.getenv("TOKEN") or os.getenv("TELEGRAM_BOT_TOKEN")
GOOGLE_DRIVE_URL = os.getenv("GOOGLE_DRIVE_URL")  # share/export link
GOOGLE_DRIVE_FILE_ID = os.getenv("GOOGLE_DRIVE_FILE_ID")  # optional explicit file id
RENDER_URL = os.getenv("RENDER_URL")
PORT = int(os.environ.get("PORT", 10000))
WEBHOOK_PATH = "webhook"

# Retry config
MAX_RETRIES = int(os.getenv("DOWNLOAD_MAX_RETRIES", "4"))
BACKOFF_FACTOR = float(os.getenv("DOWNLOAD_BACKOFF_FACTOR", "1.5"))

if not TOKEN:
    raise RuntimeError("Missing TELEGRAM token. Set environment variable TOKEN.")
if not RENDER_URL:
    raise RuntimeError("Missing RENDER_URL environment variable.")

# ====== Global state ======
user_data = {}

# ====== Flask app ======
app = Flask(__name__)

@app.route("/", methods=["GET", "HEAD"])
def health():
    return "Bot is alive!"

@app.route(f"/{WEBHOOK_PATH}", methods=["POST"])
def webhook_receiver():
    try:
        data = request.get_json(force=True)
    except Exception as e:
        print("Invalid JSON in webhook POST:", e)
        abort(400)

    if "message" not in data:
        return jsonify({"ok": True, "note": "no message"}), 200

    try:
        handle_update_sync(data)
    except Exception as e:
        print("Error handling update:", e)
    return jsonify({"ok": True}), 200

# ====== Helpers for Google Sheets / Drive ======
def _extract_drive_file_id(url: str):
    if not url:
        return None
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    m = re.search(r"id=([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    return None

def _is_google_sheets_url(url: str):
    return bool(url and "docs.google.com/spreadsheets" in url)

def _build_sheets_export_url(url: str, format="xlsx"):
    fid = _extract_drive_file_id(url)
    if not fid:
        return None
    return f"https://docs.google.com/spreadsheets/d/{fid}/export?format={format}"

def _get_confirm_token_from_response(resp_text: str):
    m = re.search(r"confirm=([0-9A-Za-z_]+)&", resp_text)
    if m:
        return m.group(1)
    m2 = re.search(r"confirm=([0-9A-Za-z_]+)\\u0026", resp_text)
    if m2:
        return m2.group(1)
    return None

def _download_with_confirm_and_retries(dl_url: str):
    session = requests.Session()
    headers = {"User-Agent": "python-requests/2.x"}
    attempt = 0
    last_exc = None

    while attempt < MAX_RETRIES:
        try:
            attempt += 1
            print(f"[Drive] download attempt {attempt} for {dl_url}")
            resp = session.get(dl_url, headers=headers, allow_redirects=True, timeout=30)
            content_type = resp.headers.get("Content-Type", "")
            status = resp.status_code
            print(f"[Drive] status {status} content-type {content_type}")
            if status == 200 and "text/html" not in content_type.lower():
                return resp.content

            text = resp.text or ""
            token = _get_confirm_token_from_response(text)
            if token:
                fid = _extract_drive_file_id(dl_url) or _extract_drive_file_id(GOOGLE_DRIVE_URL or "")
                if not fid:
                    raise RuntimeError("Cannot determine file id for confirm flow.")
                confirm_url = f"https://drive.google.com/uc?export=download&id={fid}&confirm={token}"
                print(f"[Drive] retrying with confirm token, url: {confirm_url}")
                resp2 = session.get(confirm_url, headers=headers, allow_redirects=True, timeout=30)
                resp2.raise_for_status()
                ct2 = resp2.headers.get("Content-Type", "")
                print(f"[Drive] confirm response status {resp2.status_code} content-type {ct2}")
                if "text/html" in ct2.lower():
                    raise RuntimeError("Google returned HTML even after confirm token.")
                return resp2.content

            raise RuntimeError(f"Google returned HTML page (status {status}).")

        except Exception as e:
            last_exc = e
            wait = BACKOFF_FACTOR ** attempt
            print(f"[Drive] attempt {attempt} failed: {repr(e)}; retrying in {wait:.1f}s")
            time.sleep(wait)

    raise RuntimeError(f"Download failed after {MAX_RETRIES} attempts. Last error: {repr(last_exc)}")

def _download_via_direct_link(url: str):
    fid = _extract_drive_file_id(url)
    if fid:
        dl = f"https://drive.google.com/uc?export=download&id={fid}"
    else:
        dl = url
    print(f"[Drive] Trying direct download URL: {dl}")
    content = _download_with_confirm_and_retries(dl)
    return content

def load_excel_from_google_drive(sheet_name=" "):
    last_exc = None

    # If it's a Google Sheets URL, convert to export xlsx
    if _is_google_sheets_url(GOOGLE_DRIVE_URL):
        export_url = _build_sheets_export_url(GOOGLE_DRIVE_URL, format="xlsx")
        if export_url:
            try:
                content = _download_with_confirm_and_retries(export_url)
                print("[Drive] downloaded via Google Sheets export URL")
                return pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=None)
            except Exception as e:
                last_exc = e
                print("[Drive] sheets export download failed:", repr(e))

    # 1) Try direct download if URL provided (Drive file)
    if GOOGLE_DRIVE_URL:
        try:
            content = _download_via_direct_link(GOOGLE_DRIVE_URL)
            print("[Drive] downloaded via direct link")
            return pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=None)
        except Exception as e:
            last_exc = e
            print("[Drive] direct download failed:", repr(e))

    # 2) Try explicit file id with uc link
    fid = GOOGLE_DRIVE_FILE_ID or _extract_drive_file_id(GOOGLE_DRIVE_URL or "")
    if fid:
        try:
            dl = f"https://drive.google.com/uc?export=download&id={fid}"
            content = _download_with_confirm_and_retries(dl)
            print("[Drive] downloaded via explicit file id")
            return pd.read_excel(io.BytesIO(content), sheet_name=sheet_name, header=None)
        except Exception as e:
            last_exc = e
            print("[Drive] explicit file id download failed:", repr(e))

    guidance = (
        "Failed to download Excel from Google Drive. Last error: {}\n"
        "Possible causes:\n"
        "- The share link is not set to Anyone with the link.\n"
        "- Google returned an HTML page (virus scan or preview) instead of file.\n"
        "- The link is not a direct download link or Google throttled requests.\n\n"
        "Fix options:\n"
        "1) Ensure the file is shared 'Anyone with the link' and set GOOGLE_DRIVE_URL to the export link: https://docs.google.com/spreadsheets/d/<FILEID>/export?format=xlsx\n"
        "2) Provide GOOGLE_DRIVE_FILE_ID and ensure the file is shared.\n"
    ).format(last_exc)
    raise RuntimeError(guidance)

# ====== Logging user queries ======
def save_log(user_id, name, company, question, timestamp):
    try:
        try:
            wb = load_workbook(LOG_FILE)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["User ID", "Tên", "Công ty", "Câu hỏi", "Thời gian"])
        ws.append([user_id, name, company, question, timestamp])
        wb.save(LOG_FILE)
    except Exception as e:
        print("❌ Lỗi ghi log:", e)

# ====== Telegram helper ======
TELEGRAM_API = f"https://api.telegram.org/bot{TOKEN}"

def send_message(chat_id: int, text: str, parse_mode: str = None):
    payload = {"chat_id": chat_id, "text": text}
    if parse_mode:
        payload["parse_mode"] = parse_mode
        payload["disable_web_page_preview"] = True
    try:
        resp = requests.post(f"{TELEGRAM_API}/sendMessage", data=payload, timeout=15)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print("Failed to send message:", e, resp.text if 'resp' in locals() else "")
        return None

# ====== Core handling ======
def _html_link(url: str, label: str):
    # Simple helper to produce safe HTML anchor
    return f"<a href=\"{url}\">{label}</a>"

def handle_update_sync(update_json: dict):
    msg = update_json.get("message", {})
    if not msg:
        return

    chat = msg.get("chat", {})
    from_user = msg.get("from", {})
    chat_id = chat.get("id")
    user_id = from_user.get("id")
    text = msg.get("text", "").strip() if msg.get("text") else ""

    if not chat_id or not user_id:
        return

    if text.startswith("/start"):
        user_data[user_id] = {"step": "name"}
        send_message(chat_id, "Xin chào! Vui lòng nhập Tên của bạn:")
        return

    if text.startswith("/help"):
        help_text = (
            "📖 Hướng dẫn sử dụng bot:\n\n"
            "- /start: Bắt đầu.\n"
            "- /list_dest: Liệt kê mã Dest.\n"
            "- /help: Hiển thị hướng dẫn."
        )
        send_message(chat_id, help_text)
        return

    if text.startswith("/list_dest"):
        try:
            df = load_excel_from_google_drive(sheet_name=" ")
            dest_values = df[1].dropna().unique()
            dest_list = ", ".join(sorted(dest_values.astype(str)))
            answer = f"📋 Danh sách tất cả Dest trong cột B:\n{dest_list}"
        except Exception as e:
            answer = f"⚠️ Có lỗi xảy ra khi đọc file: {e}"
            print("[Error] list_dest:", e)
        # list_dest has no HTML links, plain send is fine
        send_message(chat_id, answer)
        return

    state = user_data.get(user_id, {})
    step = state.get("step")

    if step == "name":
        user_data[user_id]["name"] = text
        user_data[user_id]["step"] = "company"
        send_message(chat_id, "Cảm ơn! Bây giờ hãy nhập Tên công ty:")
        return

    if step == "company":
        user_data[user_id]["company"] = text
        user_data[user_id]["step"] = "done"
        send_message(chat_id, "✅ Đã lưu thông tin. Giờ bạn có thể nhập mã Dest để tra cứu.")
        return

    if step != "done":
        send_message(chat_id, "⚠️ Vui lòng nhập Tên và Công ty trước bằng lệnh /start.")
        return

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_log(user_id, user_data[user_id].get("name", ""), user_data[user_id].get("company", ""), text, timestamp)

    try:
        df = load_excel_from_google_drive(sheet_name=" ")
        row = df[df[1].astype(str).str.strip().str.lower() == text.lower()]

        outlook_anchor = _html_link(OUTLOOK_LINK, "Space Outlook")
        tact_anchor = _html_link(TACT_LINK, "TACT Rate")

        if not row.empty:
            r = row.iloc[0]
            c_val = f"{r[2]:.2f}" if pd.notnull(r[2]) else ""
            i_val = str(r[8]) if pd.notnull(r[8]) else ""
            j_val = str(r[9]) if pd.notnull(r[9]) else ""
            k_val = pd.to_datetime(r[10], errors="coerce")
            k_val = k_val.strftime("%d/%m/%Y") if pd.notnull(k_val) else ""
            l_val = pd.to_datetime(r[11], errors="coerce")
            l_val = l_val.strftime("%d/%m/%Y") if pd.notnull(l_val) else ""

            extra_text = "\n".join(
                str(df.iloc[i, 13]) for i in range(1, 6) if pd.notnull(df.iloc[i, 13])
            )

            answer = (
                f"📊 Kết quả cho Dest: {text.upper()}\n"
                f"- Giá All-in USD/kg: {c_val}\n"
                f"- Điều kiện 1: {i_val}\n"
                f"- Điều kiện 2: {j_val}\n"
                f"- Valid from: {k_val}\n"
                f"- Valid till: {l_val}\n\n"
                f"📌 Thông tin bổ sung:\n{extra_text}\n\n"
                f"🔗 {outlook_anchor}\n"
                f"🔗 {tact_anchor}"
            )
            # send with HTML parse mode so anchors render
            send_message(chat_id, answer, parse_mode="HTML")
            return
        else:
            answer = (
                "Xin lỗi, chưa có dữ liệu cho giá trị này.\n"
                "👉 Bạn có thể dùng lệnh /list_dest để xem danh sách Dest có sẵn.\n\n"
                f"🔗 {outlook_anchor}\n"
                f"🔗 {tact_anchor}"
            )
            send_message(chat_id, answer, parse_mode="HTML")
            return

    except Exception as e:
        answer = f"⚠️ Có lỗi xảy ra khi tra cứu: {e}"
        print("[Error] lookup:", e)
        send_message(chat_id, answer)
        return

# ====== Webhook setup helper ======
def set_telegram_webhook(webhook_url: str):
    url = f"https://api.telegram.org/bot{TOKEN}/setWebhook"
    resp = requests.post(url, data={"url": webhook_url}, timeout=15)
    resp.raise_for_status()
    j = resp.json()
    if not j.get("ok"):
        raise RuntimeError(f"setWebhook returned not ok: {j}")
    return j

# ====== Entrypoint ======
if __name__ == "__main__":
    webhook_url = RENDER_URL.rstrip("/") + f"/{WEBHOOK_PATH}"
    print("Webhook URL will be:", webhook_url)
    try:
        res = set_telegram_webhook(webhook_url)
        print("✅ Telegram webhook set to:", webhook_url)
    except Exception as e:
        print("⚠️ setWebhook failed (will still run):", e)
    print("Starting Flask on port", PORT)
    app.run(host="0.0.0.0", port=PORT)
